import openpyxl

wb = openpyxl.load_workbook('/Users/apple/Downloads/ยอดขาย 1-17.xlsx', data_only=True)
sheet = wb['Sheet1']

print("--- Dump of All Rows in Sheet1 ---")
for r in range(1, sheet.max_row + 1):
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    # Filter out rows that are entirely empty
    if any(v is not None for v in row_vals):
        row_str = [str(v) if v is not None else "" for v in row_vals]
        # Print row index, and first few values (especially columns 8, 9, 40 which have details)
        print(f"Row {r:02d}: Col8={row_str[7]:15s} Col9={row_str[8]:20s} Total={row_str[-1]:10s}")
