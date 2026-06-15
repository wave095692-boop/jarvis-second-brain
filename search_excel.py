import openpyxl

wb = openpyxl.load_workbook('/Users/apple/Downloads/ยอดขาย 1-17.xlsx', data_only=True)
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"Searching sheet: {sheet_name}")
    for r in range(1, sheet.max_row + 1):
        for c in range(1, sheet.max_column + 1):
            val = sheet.cell(r, c).value
            if val is not None and "095692" in str(val):
                print(f"Found in Row {r}, Col {c}: {val}")
